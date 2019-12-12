# -*- coding: UTF-8 -*-
#__author__ = 'Luke'
#获取链家指定小区的二手房信息
import re,os,sys
import time
import random
import requests
import openpyxl
from bs4 import BeautifulSoup

os.chdir(os.path.dirname(sys.argv[0]))      #更改脚本运行路径到脚本所在目录


xiaoqu_dict = {'龙园':4511059517137,'新悦城':457350664310735,'逸都国际':457351702924987,'幸福里':456945018753103,'万象新天':456942057574347}

hds=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
    {'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
    {'User-Agent':'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'},\
    {'User-Agent':'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:34.0) Gecko/20100101 Firefox/34.0'},\
    {'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/44.0.2403.89 Chrome/44.0.2403.89 Safari/537.36'},\
    {'User-Agent':'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_8; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50'},\
    {'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50'},\
    {'User-Agent':'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0'},\
    {'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.6; rv:2.0.1) Gecko/20100101 Firefox/4.0.1'},\
    {'User-Agent':'Mozilla/5.0 (Windows NT 6.1; rv:2.0.1) Gecko/20100101 Firefox/4.0.1'},\
    {'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_0) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11'},\
    {'User-Agent':'Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; en) Presto/2.8.131 Version/11.11'},\
    {'User-Agent':'Opera/9.80 (Windows NT 6.1; U; en) Presto/2.8.131 Version/11.11'}]



def get_xiaoqu_info(xiaoqu_dict):
    info_list = []
    for xiaoqu in xiaoqu_dict:
    
        url = 'http://jn.lianjia.com/ershoufang/{}c{}/'
        start_url = url.format('',xiaoqu_dict[xiaoqu])
        html = requests.get(start_url, headers=hds[random.randint(0,len(hds)-1)])
        html_txt = html.text
        soup = BeautifulSoup(html_txt,'lxml')
        
        taoshu = soup.select('#content > div:nth-of-type(1) > div:nth-of-type(2) > h2 > span')        #总套数
        taoshuint = int(taoshu[0].string)
        taoshu_dict[xiaoqu] = taoshuint
        
        try:
            yeshu = soup.select('#content > div:nth-of-type(1) > div:nth-of-type(8) > div:nth-of-type(2) > div:nth-of-type(1)')       #总页数
            yeshuint = eval(yeshu[0]["page-data"])["totalPage"]
        except:
            yeshuint = 1    # 没有多页，只有1页
        for i in range(yeshuint):
            if i+1 == 1 :
                info_list += get_page_info(xiaoqu,soup)
            else:
                url = url.format('pg{}'.format(i+1),xiaoqu_dict[xiaoqu])
                html = requests.get(url, headers=hds[random.randint(0,len(hds)-1)])
                html_txt = html.text
                soup = BeautifulSoup(html_txt,'lxml')
                info_list += get_page_info(xiaoqu,soup)
        
        print (xiaoqu,"已完成")
    
    return taoshu_dict,info_list


def get_page_info(xiaoqu,soup):
    page_info_list = []
    infos = soup.select('#content > div:nth-of-type(1) > div:nth-of-type(6) > div')
    for eachinfo in infos:
        info1 = float((eachinfo.select('a:nth-of-type(1) > div:nth-of-type(4) > span')[0]).string)        #总价
        info2 = re.sub('\s','_',str((eachinfo.select('a:nth-of-type(2)')[0]).string))        #标题
        info345 = eachinfo.find(class_='info')
        info345re = re.sub(r'<.*?>', '',str(info345))
        info345relist = info345re.split("/")
        info3 = info345relist[1]    #布局
        info4 = float(re.sub('平米','',info345relist[2]))    #平米
        info5 = info345relist[4]    #装修
        info6 = float('%.2f'%(info1/info4))
        page_info_list.append([xiaoqu,info2,info3,info5,info4,info1,info6])     #小区名、标题、布局、装修、平米、总价(万)、单价(万)
        # print (xiaoqu,info2,info3,info5,info4,info1,info6)
    return page_info_list


def get_min_price(info_list):
    # print (info_list);input()
    min_price_dict = {}
    for xiaoqu in xiaoqu_dict:
        min_price = 123456      #初始单价12万
        for i,info in enumerate(info_list):
            if info_list[i][0] == xiaoqu and info_list[i][4] < 120 and info_list[i][6] < min_price:
                min_price = info_list[i][6]
        min_price_dict[xiaoqu] = min_price
    return min_price_dict


def write_excel(min_price_dict,taoshu_dict):      #输入二手房单价最低信息，小区套数信息
    
    excel_name = '链家信息收集最低和套数.xlsx'
    sheet_name = '链家'
    
    t = time.localtime()
    timestr = '{}月{}日'.format(t.tm_mon,t.tm_mday)
    
    seq = ['时间','龙园','新悦城','幸福里','逸都国际','万象新天']
    
    if not os.path.exists(excel_name):
        excel = openpyxl.Workbook()
        sheet = excel.create_sheet(sheet_name,0)
        excel.remove(excel['Sheet'])
        sheet.cell(row=1, column=1).value = '_______________120平以下最低单价(万/m2)________________'
        sheet.cell(row=1, column=8).value = '____________________小区总套数(套)____________________'
        for i,each in enumerate(seq):
            sheet.cell(row=2, column=i+1).value = each
            sheet.cell(row=2, column=i+1).font = openpyxl.styles.Font(bold=True)
            sheet.cell(row=2, column=i+1).fill = openpyxl.styles.PatternFill('solid', fgColor='8DB4E2')
            sheet.cell(row=2, column=i+8).value = each
            sheet.cell(row=2, column=i+8).font = openpyxl.styles.Font(bold=True)
            sheet.cell(row=2, column=i+8).fill = openpyxl.styles.PatternFill('solid', fgColor='8DB4E2')
        print ("本地没有Excel信息，已新建<{}> 。".format(excel_name))
    else:
        excel = openpyxl.load_workbook(excel_name)
        sheet = excel['链家']
        
    if sheet_name not in excel.sheetnames:
        print ("sheet名字错误")
        sys.exit()
        
    rowmax = sheet.max_row
    
    sheet.cell(row=rowmax+1, column=1).value = timestr
    sheet.cell(row=rowmax+1, column=8).value = timestr
    
    for x,colxxx in enumerate(seq[1:]):     #['龙园','新悦城','幸福里','逸都国际','万象新天']
        sheet.cell(row=1+rowmax, column=x+1+1).value = min_price_dict[colxxx]
        sheet.cell(row=1+rowmax, column=x+1+8).value = taoshu_dict[colxxx]
        
    excel.save(excel_name)
    print ("数据已保存到Excel")
    os.system('\"'+excel_name+'\"')
    return 


def yanzheng():
    try:
        excel = openpyxl.load_workbook('链家信息收集最低和套数.xlsx')
        sheet = excel['链家']
        rowmax = sheet.max_row
        t = time.localtime()
        timestr = '{}月{}日'.format(t.tm_mon,t.tm_mday)
        if timestr == sheet.cell(row=rowmax, column=1).value:
            print ("今日链家信息已经收集！")
            return True
        else:
            return False
    except:
        return False

if __name__ == '__main__':
    taoshu_dict = {}
    info_list = []
    min_price_dict = {}
    
    if yanzheng():sys.exit()
    
    taoshu_dict,info_list = get_xiaoqu_info(xiaoqu_dict)       #获取套数表、信息表
    
    min_price_dict = get_min_price(info_list)       #信息表→最低价表
    
        
    write_excel(min_price_dict,taoshu_dict)
    sys.exit()
































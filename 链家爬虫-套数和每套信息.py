# -*- coding: UTF-8 -*-
#__author__ = 'Luke'
#获取链家指定小区的二手房信息
import re,os,sys
import time
import random
import requests
import openpyxl
from bs4 import BeautifulSoup

os.chdir(os.path.dirname(sys.argv[0]))      #更改脚本运行路径到脚本所在目录


dict = {'龙园':4511059517137,'新悦城':457350664310735,'逸都国际':457351702924987,'幸福里':456945018753103,'万象新天':456942057574347}

hds=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
    {'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
    {'User-Agent':'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'},\
    {'User-Agent':'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:34.0) Gecko/20100101 Firefox/34.0'},\
    {'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/44.0.2403.89 Chrome/44.0.2403.89 Safari/537.36'},\
    {'User-Agent':'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_8; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50'},\
    {'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50'},\
    {'User-Agent':'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0'},\
    {'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.6; rv:2.0.1) Gecko/20100101 Firefox/4.0.1'},\
    {'User-Agent':'Mozilla/5.0 (Windows NT 6.1; rv:2.0.1) Gecko/20100101 Firefox/4.0.1'},\
    {'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_0) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11'},\
    {'User-Agent':'Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; en) Presto/2.8.131 Version/11.11'},\
    {'User-Agent':'Opera/9.80 (Windows NT 6.1; U; en) Presto/2.8.131 Version/11.11'}]



def get_xiaoqu_info(xiaoqu):
    url = 'http://jn.lianjia.com/ershoufang/{}c{}/'
    start_url = url.format('',dict[xiaoqu])
    html = requests.get(start_url, headers=hds[random.randint(0,len(hds)-1)])
    html_txt = html.text
    soup = BeautifulSoup(html_txt,'lxml')
    
    taoshu = soup.select('#content > div:nth-of-type(1) > div:nth-of-type(2) > h2 > span')        #总套数
    taoshuint = int(taoshu[0].string)
    taoshu_dict[xiaoqu] = taoshuint
    
    try:
        yeshu = soup.select('#content > div:nth-of-type(1) > div:nth-of-type(8) > div:nth-of-type(2) > div:nth-of-type(1)')       #总页数
        yeshuint = eval(yeshu[0]["page-data"])["totalPage"]
    except:
        yeshuint = 1    # 没有多页，只有1页
    for i in range(yeshuint):
        if i+1 == 1 :
            get_info(xiaoqu,soup)
        else:
            url = url.format('pg{}'.format(i+1),dict[xiaoqu])
            html = requests.get(url, headers=hds[random.randint(0,len(hds)-1)])
            html_txt = html.text
            soup = BeautifulSoup(html_txt,'lxml')
            get_info(xiaoqu,soup)


def get_info(xiaoqu,soup):
    infos = soup.select('#content > div:nth-of-type(1) > div:nth-of-type(6) > div')
    for eachinfo in infos:
        info1 = float((eachinfo.select('a:nth-of-type(1) > div:nth-of-type(4) > span')[0]).string)        #总价
        info2 = re.sub('\s','_',str((eachinfo.select('a:nth-of-type(2)')[0]).string))        #标题
        info345 = eachinfo.find(class_='info')
        info345re = re.sub(r'<.*?>', '',str(info345))
        info345relist = info345re.split("/")
        info3 = info345relist[1]    #布局
        info4 = float(re.sub('平米','',info345relist[2]))    #平米
        if info4 > 120:continue
        info5 = info345relist[4]    #装修
        info6 = float('%.2f'%(info1/info4))
        info_list.append([xiaoqu,info2,info3,info5,info4,info1,info6])     #小区名、标题、布局、装修、平米、总价(万)、单价(万)
        # print (xiaoqu,info2,info3,info5,info4,info1,info6)



def get_min_price(dict,info_list):
    min_price_dict = {}
    for xiaoqu in dict:
        min_price = 100000
        for i,info in enumerate(info_list):
            if info_list[i][0] == xiaoqu and info_list[i][5] < min_price:
                min_price = info_list[i][5]
        min_price_dict[xiaoqu] = min_price
    return min_price_dict





def write_excel(taoshu_dict,info_list):      #输入小区套数信息，二手房详细信息    #新建一个sheet，放在最前
    excel_name = r'链家信息收集.xlsx'
    sheet_name = time.strftime('%m%d',time.localtime())
    
    if not os.path.exists(excel_name):
        excel = openpyxl.Workbook()
        print ("本地没有Excel信息，已新建<{}> 。".format(excel_name))
    else:
        excel = openpyxl.load_workbook(r'链家信息收集.xlsx')
        if sheet_name in excel.sheetnames:
            print ("今日信息已经收集！")
            sys.exit()
    
    sheet = excel.create_sheet(sheet_name,0)
    sheet.cell(row=1, column=1).value = '小区'
    sheet.cell(row=1, column=2).value = '总套数'
    for i,p in enumerate(taoshu_dict):                                                                  #写入套数字典到sheet
        sheet.cell(row=i+2, column=1).value = p
        sheet.cell(row=i+2, column=2).value = taoshu_dict[p]
    
    sheet_top_row = ['小区名','标题','布局','装修','平米','总价(万)','单价(万)']
    row_first = len(taoshu_dict)+5
    for j,q in enumerate(sheet_top_row):                                                                     #写入标题到sheet，默认1行标题
        sheet.cell(row=row_first, column=j+1).value = q
        sheet.cell(row=row_first, column=j+1).font = openpyxl.styles.Font(bold=True)                                               #加粗
        sheet.cell(row=row_first, column=j+1).fill = openpyxl.styles.PatternFill('solid', fgColor='8DB4E2')                        #设置填充颜色
    
    for i,p in enumerate(info_list):                                                                  #写入主要内容到sheet
        for j,q in enumerate(p):
            sheet.cell(row=i+1+row_first, column=j+1).value = q
        
    sheet.column_dimensions['B'].width = 40
    
    # min_price_dict = get_min_price(dict,info_list)            #汇总的部分 暂时不写
    # sheet = excel["汇总"]
    # row_first = sheet.max_row
    # for i,p in enumerate(min_price_dict):
        # sheet.cell(row=row_first+1, column=i+1).value = min_price_dict[p]
        
    excel.save(excel_name)
    print ("数据已保存到Excel")
    os.system('\"'+excel_name+'\"')
    return 


def main():
    global taoshu_dict
    global info_list
    taoshu_dict = {}
    info_list = []
    for xiaoqu in dict:
        get_xiaoqu_info(xiaoqu)
        print (xiaoqu,"已完成")
    # get_xiaoqu_info('万象新天')
    write_excel(taoshu_dict,info_list)
    sys.exit()

if __name__ == '__main__':
    main()































